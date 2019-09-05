import random
import time

def not_duplicate_generator():
    #重複チェック用フラグ
    flag =1
    #重複チェック通るまでループ
    while flag==1:
        #数字生成
        num=random.randint(12,987)
        num=str(num)
        #0パディング、しないとstrがout of range
        num=num.zfill(3)
        #桁で分解
        numa=num[0]#百の位
        numb=num[1]#十の位
        numc=num[2]#一の位
        #重複チェック
        if numa!=numb and numb!=numc and numc!=numa:
            flag=0
    return num

def judge(prev,next):
    prev=prev.zfill(3)
    next=next.zfill(3)
    #桁で分解
    nexta=next[0]#百の位
    nextb=next[1]#十の位
    nextc=next[2]#一の位
    #桁で分解
    preva=prev[0]#百の位
    prevb=prev[1]#十の位
    prevc=prev[2]#一の位
    eat=0
    if preva==nexta:
        eat+=1
    if prevb==nextb:
        eat+=1
    if prevc==nextc:
        eat+=1
    bite=0
    if preva==nextb or preva==nextc:
        bite+=1
    if prevb==nexta or prevb==nextc:
        bite+=1
    if prevc==nexta or prevc==nextb:
        bite+=1
    same=0
    if preva==nexta or preva==nextb or preva==nextc:
        same+=1
    if prevb==nexta or prevb==nextb or prevb==nextc:
        same+=1
    if prevc==nexta or prevc==nextb or prevc==nextc:
        same+=1
    return eat,bite,same

def strategy(prevcom,com):
    redoflag=0
    while redoflag==0:
        redoflag=0
        com=not_duplicate_generator()
        eat,bite,same=judge(prevcom,com)
        if eat == 3:
            break
        if eat+bite==1 and same > 2:
            pass
        elif eat+bite==1 and same == 0:
            pass
        elif eat+bite==2 and same == 3:
            pass
        elif eat+bite==2 and same < 2:
            pass
        elif eat+bite==3 and same == 0:
            pass
        elif bite==0 and eat == 0:
            pass
        elif eat==0 and eat > 0:
            pass
        elif eat==1 and eat > 1:
            pass
        else:
            redoflag=1
    return com

def one_point_four():
    timesfour=0
    tfour=time.time()
    import random
    player=not_duplicate_generator()
    i=0
    #com初回call
    com=not_duplicate_generator()
    #eat,bite判定
    eat,bite,same=judge(player,com)
    while eat!=3:
        prevcom=com
        com=strategy(prevcom,com)
        eat,bite,same=judge(player,com)
        timesfour+=1
    tfour=time.time()-tfour
    return tfour,timesfour

import openpyxl as px
wb = px.Workbook()
name=input("bookname\n")
vs="1.4_"
name=vs+name+".xlsx"
wb.save(name)
ws=wb.active
ws.cell(row=1,column=1,value="times(ver1.4)")
ws.cell(row=1,column=2,value="time(ver1.4)")

ceil=int(input("回数\n"))

row_num=2
count=0
while count < ceil:
    tfour,timesfour=one_point_four()
    ws.cell(row=row_num,column=1,value=timesfour)
    ws.cell(row=row_num,column=2,value=tfour)
    row_num+=1
    count+=1
    if count % 10000 ==0:
        print(count)
        print("try end")

wb.save(name)
input("operation succeeded")

#prevcomとplayerのjudge結果を見て、strategyに適したcomを出さなきゃいけない