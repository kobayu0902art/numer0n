import time
import random
def not_duplicate_generator():
    #重複チェック用フラグ
    flag =1
    #重複チェック通るまでループ
    while flag==1:
        #数字生成
        num=random.randint(12,987)
        num=str(num)
        #0パディング、しないとstrがout of range
        num=num.zfill(3)
        #桁で分解
        numa=num[0]#百の位
        numb=num[1]#十の位
        numc=num[2]#一の位
        #重複チェック
        if numa!=numb and numb!=numc and numc!=numa:
            flag=0
    return num
def judge(prev,next):
    prev=prev.zfill(3)
    next=next.zfill(3)
    #桁で分解
    nexta=next[0]#百の位
    nextb=next[1]#十の位
    nextc=next[2]#一の位
    #桁で分解
    preva=prev[0]#百の位
    prevb=prev[1]#十の位
    prevc=prev[2]#一の位
    eat=0
    if preva==nexta:
        eat+=1
    if prevb==nextb:
        eat+=1
    if prevc==nextc:
        eat+=1
    bite=0
    if preva==nextb or preva==nextc:
        bite+=1
    if prevb==nexta or prevb==nextc:
        bite+=1
    if prevc==nexta or prevc==nextb:
        bite+=1
    same=0
    if preva==nexta or preva==nextb or preva==nextc:
        same+=1
    if prevb==nexta or prevb==nextb or prevb==nextc:
        same+=1
    if prevc==nexta or prevc==nextb or prevc==nextc:
        same+=1
    return eat,bite,same

def strategy(player):
    redoflag=0
    while redoflag==0:
        redoflag=0
        com=not_duplicate_generator()
        eat,bite,same=judge(player,com)
        if eat == 3:
            break
        if eat+bite==1 and same > 2:
            pass
        elif eat+bite==1 and same == 0:
            pass
        elif eat+bite==2 and same == 3:
            pass
        elif eat+bite==2 and same < 2:
            pass
        elif eat+bite==3 and same == 0:
            pass
        elif bite==0 and eat == 0:
            pass
        elif eat==0 and eat > 0:
            pass
        elif eat==1 and eat > 1:
            pass
        else:
            redoflag=1
    return com
def one_point_one():
    import random
    #重複チェック用フラグ
    replayflag =1
    #重複チェック通るまでループ
    while replayflag==1:
        #数字生成
        num=random.randint(12,987)
        num=str(num)
        #0パディング、しないとstrがout of range
        num=num.zfill(3)
        #桁で分解
        a=num[0]#百の位
        b=num[1]#十の位
        c=num[2]#一の位
        #重複チェック
        if a!=b and b!=c and c!=a:
            replayflag=0
    eat=0
    tone=time.time()
    timesone=0 
    while eat!=3:
        import random
        #重複チェック用フラグ
        replayflag =1
        #重複チェック通るまでループ
        while replayflag==1:
            #コール用数字生成
            call=random.randint(12,987)
            call=str(call)
            #0パディング、しないとstrがout of range
            call=call.zfill(3)
            #桁で分解
            calla=call[0]#百の位
            callb=call[1]#十の位
            callc=call[2]#一の位
            #重複チェック
            if calla!=callb and callb!=callc and callc!=calla:
                replayflag=0
        #eatチェック
        eat=0
        if a==calla:
            eat+=1
        if b==callb:
            eat+=1
        if c==callc:
            eat+=1
        #biteチェック
        bite=0
        if a==callb or a==callc:
            bite+=1
        if b==calla or b==callc:
            bite+=1
        if c==calla or c==callb:
            bite+=1
        timesone+=1   
    tone=time.time()-tone
    return tone,timesone

def one_point_three():
    timesthree=0
    tthree=time.time()
    import random
    
    player=not_duplicate_generator()
    eat=0 
    timesthree=0
    tthree=time.time()
    while eat!=3:
        if eat == 3:
            break
        com=strategy(player)
        eat,bite,same=judge(player,com)
        timesthree+=1
    tthree=time.time()-tthree
    return tthree,timesthree

import openpyxl as px
wb = px.Workbook()
name=input("bookname\n")
vs="1.1vs1.3_"
name=vs+name+".xlsx"
wb.save(name)
ws=wb.active
ws.cell(row=1,column=1,value="times(ver1.1)")
ws.cell(row=1,column=2,value="time(ver1.1)")
ws.cell(row=1,column=5,value="times(ver1.3)")
ws.cell(row=1,column=6,value="time(ver1.3)")

ceil=int(input("回数\n"))

row_num=2
count=0
while count < ceil:
    onetime,onetimes=one_point_one()
    ws.cell(row=row_num,column=1,value=onetimes)
    ws.cell(row=row_num,column=2,value=onetime)
    threetime,threetimes=one_point_three()
    ws.cell(row=row_num,column=5,value=threetimes)
    ws.cell(row=row_num,column=6,value=threetime)
    row_num+=1
    count+=1

wb.save(name)
input("operation succeeded")